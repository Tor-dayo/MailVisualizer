from openpyxl import Workbook
from openpyxl.styles import Font


def export_excel(mails, filename):

    wb = Workbook()
    ws = wb.active
    ws.title = "Mail List"

    headers = [
        "No",
        "日時",
        "差出人",
        "宛先",
        "件名",
        "本文プレビュー",
        "Message-ID"
    ]

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.value = h
        cell.font = Font(bold=True)

    for r, mail in enumerate(mails, start=2):

        ws.cell(r,1).value = mail["no"]
        ws.cell(r,2).value = mail["date"]
        ws.cell(r,3).value = mail["from"]
        ws.cell(r,4).value = mail["to"]
        ws.cell(r,5).value = mail["subject"]
        ws.cell(r,6).value = mail["preview"]
        ws.cell(r,7).value = mail["message_id"]

    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 80

    wb.save(filename)