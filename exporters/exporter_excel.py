import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


def clean_excel(value, limit=32000):
    if value is None:
        return ""

    value = str(value)
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    value = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)

    return value[:limit]


def export_excel(mails, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mail List"

    headers = [
        "No",
        "日時",
        "差出人",
        "宛先",
        "件名",
        "本文プレビュー",
        "Message-ID",
    ]

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.value = h
        cell.font = Font(bold=True)

    for r, mail in enumerate(mails, start=2):
        ws.cell(r, 1).value = clean_excel(mail.no)
        ws.cell(r, 2).value = clean_excel(mail.date)
        ws.cell(r, 3).value = clean_excel(mail.sender)
        ws.cell(r, 4).value = clean_excel(mail.to)
        ws.cell(r, 5).value = clean_excel(mail.subject)
        ws.cell(r, 6).value = clean_excel(mail.preview, 1000)
        ws.cell(r, 7).value = clean_excel(mail.message_id)

        ws.cell(r, 6).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 80
    ws.column_dimensions["G"].width = 45

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)